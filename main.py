# -*- unicode:utf-8 -*-

import json
import os
import pickle
import random
import sys
import tkinter as tk
import tkinter.messagebox as msg
from multiprocessing import Pool
from time import time
from tkinter.scrolledtext import ScrolledText

from openpyxl import load_workbook  # 导入模块

UserNameFu = None
def read_pickle(filename,data=[]):
    try:
        with open(filename,'rb') as f:
            return pickle.load(f)
    except IOError as error:
        print(error)
        with open(filename,'wb') as f:
            return pickle.dump(data,f)
        read_pickle(filename,data)
def read( file):
    with open(file,'r') as F:
        return json.load(F)
def send(cmd):
    try:
        with open("../data.1",'r') as f:
            a = f.read()
        with open("../data.1",'w') as f:
            f.write(cmd)
    except:
        with open("./data.1",'w') as f:
            f.write(cmd)

class Caozuo(): #用于操作题库获取信息与写入错题
    def __init__(self):

        self.wb = load_workbook ( filename="题库.xlsx" )

        self.sheet_TiKu = self.wb["题库"]
        self.sheet_CuoTi = self.wb["类型一栏"]

    def Check( self ): #获取题库的问题总数
        for i in range(1,10000):
            name = "A%s"%(i)
            # print(self.sheet_TiKu[name].value)
            if self.sheet_TiKu[name].value == None:
                # print(i-1)
                return i-1
            else:pass
    def Check_A( self ):
        wb = load_workbook ( filename="错题集.xlsx" )
        ws = wb["错题集"]
        for i in range(1,10000):
            name = "A%s" %(i)
            if ws[name].value == None:
                return i-1

    def write_TiKu( self ,List): # 记录错题
        def main():
            wb = load_workbook ( filename="错题集.xlsx" )
            ws = wb["错题集"]
            ws["A1"] = "题干"
            ws["B1"] = 'A选项'
            ws["C1"] = 'B选项'
            ws["D1"] = "C选项"
            ws["E1"] = "D选项"
            ws["F1"] = "正确答案"
            ws["G1"] = "你的答案"
            print("记录错题中...")
            coun = self.Check_A()+1
            if coun == 1:
                coun = 2
            for L in List:

                    i = str(coun)


                    TiGan = "A" + i
                    A = "B" + i
                    B = "C" + i
                    C = "D" + i
                    D = "E" + i
                    T = "F" + i
                    Y = "G" + i
                    ws[TiGan] = L[0]
                    ws[A] = L[1]
                    ws[B] = L[2]
                    ws[C] = L[3]
                    ws[D] = L[4]
                    ws[T] = L[5]
                    ws[Y] = L[6]
                    coun += 1
            wb.save("./错题集.xlsx")
        main()



    def read_TiKu_type(self,type=None,modle="T"): # 函数有两种模式 modle="T"时  return 题库的类型 为F时 返回符合type的题
        if modle == "T":
            row = 1
            list = []
            while True:
                name = "A%s"%(row)

                if self.sheet_CuoTi[name].value == None:
                    break
                else:
                    list.append(self.sheet_CuoTi[name].value)
                    row += 1

            print(list);return list
        elif modle == "F":

            List_check = [] #c存储符合条件的单元格列数
            end = self.Check() + 1
            # print("end -- %s"%(end))

            for i in range(2,end):
                type_row = "G%s"%(i)
                if type == None:
                    List_check.append(i)
                elif type != None:
                    if self.sheet_TiKu[type_row].value == type:
                        List_check.append(i)
                    # print(i)
            if List_check == []:
                tk.messagebox.showerror('Tip',"该类型的题目数为0")
            else:
                Len = len(List_check)
                # print(List_check)
                print("已为你找到了%s道类似的题"%(Len))
            All = []

            for i in List_check:
                Tigan = "A%s" % (i)
                A = "B%s" % (i)
                B = "C%s" % (i)
                C = "D%s" % (i)
                D = "E%s" % (i)
                True_Answer = "F%s" % (i)
                type_que = "G%s" % (i)
                List = [self.sheet_TiKu[Tigan].value , self.sheet_TiKu[A].value ,
                        self.sheet_TiKu[B].value , self.sheet_TiKu[C].value ,
                        self.sheet_TiKu[D].value , self.sheet_TiKu[True_Answer].value,self.sheet_TiKu[type_que].value]
                # print ( List )
                All.append ( List )
            return All

    def Cuoti(self): # 返回所有的错题
        must = self.Check_A()

        List = []
        wb = load_workbook ( filename="错题集.xlsx" )
        sheet_CuoTi = wb["错题集"]
        for Number in range(2,must):
            Name_WenTi = "A%s" % (Number)
            Name_A = "B%s" % (Number)
            Name_B = "C%s" % (Number)
            Name_C = "D%s" % (Number)
            Name_D = "E%s" % (Number)
            Name_key = "F%s" % (Number)
            Name_you = "G%s" % (Number)
            List.append((sheet_CuoTi[Name_WenTi].value,
                sheet_CuoTi[Name_A].value,
                sheet_CuoTi[Name_B].value,
                sheet_CuoTi[Name_C].value,
                sheet_CuoTi[Name_D].value,
                sheet_CuoTi[Name_key].value,
                sheet_CuoTi[Name_you].value))
        #添加
        #去重
        Set = set(List)
        List = list(Set)
        print("已去重")
        return List
class Windows():
    def __init__(self):
        self.Frist = [True,True]
        self.font = 18
        self.ask = None
        self.start_time = 0
        self.end_time = 0
        # 考试开始与结束的时间
        self.FenShu = 0
        self.indexa = 0 # 考试的索引值
        self.mode = False #是否进入了答题模式
        self.Leavel_root_number =0
        self.c = 0 # 首次打开程序弹出的窗口，在弹出之后便不在弹出（打开程序只弹一次）
        self.Check_List = []
        self.True_Answer_type = None
        self.Update_list = None
        self.count=0
        self.change = 3
        self.Error = []
        self.mark=0
        self.geometry = True
        self.True_count = 0




        try:
            self.All = read ( "AL" )  # 考试总题数
        except FileNotFoundError as e:
            print("文件阅读错误！（无大碍报错）")
        print("已经还原为初始化状态")

        try:
            self.tool = Caozuo ( )
            self.root = tk.Tk()
            self.root.title("题库")
            self.root.iconbitmap('./main.ico')
            self.Username = UserNameFu
            self.root.geometry("250x250")
            os.chdir ( "./%s"%(UserNameFu) )

            with open("./data2.json",'r') as f:
                self.dict = json.load(f)
            self.Main()


            self.root.resizable(False,False)

            self.root.mainloop()

            sys.exit()
        except:
            pass
    def Worre( self ):
        msg.showinfo("Tips","由于宽度限制，用户可以自行调整其宽度")
        self.root.resizable(True,True)
        PZ1 = tk.Label (self.root)
        PZ1.place ( rely=0 , relx=0 , relwidth=1 , relheight=1 )
        self.root.geometry("1520x400")
        List = self.tool.Cuoti()
        print(List)
        T = ScrolledText(font=("隶书",18))
        T.place(x=60, y=30,relwidth=0.9,relheight=1)
        def add():
            self.font += 1
            T["font"] = ("隶书",self.font)
        Font_add = tk.Button(text=" A+(字体加大) ",command=add)
        def cost():
            if self.font >= 0:
                self.font -= 1
            else:
                msg.showinfo("Tip","字体太小了")
            T["font"] = ("隶书", self.font)
        Font_cosr = tk.Button(text=" A-(字体减小) ",command=cost)
        Font_add.place(rely=0,x=90)
        Font_cosr.place(rely=0,x=360)

        for i in List:
            text = "%s\n A.%s\n B.%s\n C.%s\n D.%s\n你的答案:%s\n正确答案:%s\n       " %(i[0],i[1],i[2],i[3],i[4],i[6],i[5])
            T.insert('end' , text)
            T.insert('end', ' \n')

    def setting( self ):
        PZ1 = tk.Label(self.root)
        PZ1.place(rely=0,relx=0,relwidth=1,relheight=1)
        t = tk.Label(self.root,text="出题数目：")
        t.place(relx=0, rely=0.01)
        v = tk.StringVar ( )
        v2 = tk.StringVar()
        s2 = tk.Scale (self.root,
                     from_=1 ,  # 设置最小值
                     to=100 ,  # 设置最大值
                     orient=tk.HORIZONTAL ,  # 设置横向
                     resolution=1 ,  # 设置步长
                     tickinterval=10 ,  # 设置刻度
                     length=240 ,  # 设置像素
                     variable=v )  # 绑定变量
        s2.place (relx=0,rely=0.1
                  )
        s3 = tk.Scale(self.root,
                      from_=1,  # 设置最小值
                      to=100,  # 设置最大值
                      orient=tk.HORIZONTAL,  # 设置横向
                      resolution=1,  # 设置步长
                      tickinterval=10,  # 设置刻度
                      length=240,  # 设置像素
                      variable=v2)  # 绑定变量
        s3.place(relx=0,rely=0.5)

        def S():
            self.Leavel_root_count = int(v2.get())
            self.All = int(v.get())
            if msg.askquestion("Tip",'确认保存？你需要重启才可以生效\n(目前答题界面只适用于考试模块，并无法运用于专题测试)') == 'yes':
                msg.showinfo("Tip","已将出题数修改为%s\n"%self.All)
                with open("../AL","w") as F:
                    json.dump(self.All,F)
                with open("../ALLL.json","w") as F:
                    json.dump(self.Leavel_root_count,F)
                s2.destroy()

            a.destroy()
            self.Main()


        a = tk.Button(self.root,text="保存",command=S)
        a.place(relx=0.7,rely=0)

    def delet( self ):
        if msg.askquestion("确定删除","删除是不可逆的操作，请稍加思考（按是继续）") == 'yes':
            name = read("../pass.json")
            name.pop(UserNameFu)
            self.root.destroy()
            with open("../pass.json",'w') as F:
                json.dump(name,F)
            os.chdir ( "../" )


    def Add(self,type_que,mode=1): #mode代表分数
        # type_que = Choose.get()
        if type_que == "单选":
            type_que = "单选"
        elif type_que == "多选":
            type_que = "多选"
        try:
            if mode == 1:
                self.dict[type_que] += 1
            if mode == -1:
                self.dict[type_que] -= 1
        except Exception as ex:
                print(ex)
    def Z_KaoShi(self):
        self.Error = []
        from tkinter import ttk
        PZ1 = tk.Label (self.root )
        msg.showinfo ( "Tip" , "由于宽度限制，用户可以自行调整其宽度" )
        self.root.resizable ( True , False )
        list_type = self.tool.read_TiKu_type()
        PZ1.place ( rely=0 , relx=0 , relwidth=1 , relheight=1 )
        Tip = tk.Label(self.root,text="请选择考试试卷类型：")

        Tip.place(relx=0.2,rely=0.3)
        Choose = ttk.Combobox(self.root)



        Choose["values"] = list_type

        def Start(num):

            # self.Update_list =self.tool.read_TiKu_type(modle="F", type=Choose.get())

            # self.Update_list = random.choice ( self.tool.read_TiKu_type ( modle="F" , type=Choose.get ( ) ) )
            # self.True_Answer_type = random.choice(self.Update_list[5])
            #
            # print ( ">>>%s"%(self.Update_list) )

            if Choose.get() in list_type:
                All = self.tool.read_TiKu_type(modle="F",type=Choose.get())
                #Over There Is Type Que

                Must = len(All)

                if All != []:
                    # self.Update_list=All[num]
                    # self.True_Answer_type =self.Update_list[5]
                    def randon_index():
                        print(Must)
                        random_index = random.randint(1,Must)
                        return random_index
                    def update(key):#1.检查题，加分或不加分 2.刷新界面，修改题目

                        if self.True_Answer_type == key: # 做对了
                            # print(Tiw[ans])
                            self.True_count += 1
                            self.dict['正确率'] += 2
                            self.Add(type_que=Choose.get()) # 专项得分

                            Start(num+1)

                        else:
                            self.Add(type_que=Choose.get(),mode=-1)# 专项减分
                            if self.dict['正确率']>0:self.dict['正确率'] -= 2
                            else:self.dict['正确率'] = 0
                            name = read('data1.json')
                            if self.True_count > name:
                                with open("data1.json",'w') as f:
                                    json.dump(self.True_count,f)
                                print("做到了连对 %s 道题刷新了之前连对 %s 道题的记录,\n继续做题吧"%(self.True_count,name))
                            self.True_count = 0
                            self.Error.append(self.Update_list)
                            tk.messagebox.showerror("Tip","正确答案：%s"%self.True_Answer_type)
                            self.change -= 1
                            if self.change == 0:
                                with open('./data2.json','w') as f:
                                    json.dump(self.dict,f)
                                print(self.dict)
                                chanage = tk.messagebox.askquestion ( '考试结束' , "你已经没有机会了，考试结束\n是否保存错题" )
                                # print(chanage)
                                if chanage == 'yes':

                                    self.tool.write_TiKu(self.Error)
                                self.Main()
                                print("谢谢使用")

                                self.mark = 0
                                self.change=3

                    def A():
                        update("A")

                    def D():
                        update("D")

                    def C():
                        update("C")

                    def B():
                        update("B")

                    def abort():
                        self.tool.write_TiKu(self.Error)
                        self.Main()
                        print("谢谢使用")

                    def jump(jumpnum):
                        Start(jumpnum)


                    PZ1 = tk.Label (self.root )
                    PZ1.place ( rely=0 , relx=0 , relwidth=1 , relheight=1 )
                    self.root.geometry ( "400x400" )

                    TiGan = tk.Text ( self.root)

                    TiGan.place ( relx=0.15 , rely=0.1 , relwidth=0.7 , relheight=0.2 )
                    TiGan.insert ( 'end' , self.Update_list[0] )
                    if len(self.Update_list[5]) == 1:
                        A = tk.Button (self.root, command=A , text="A.%s" % (self.Update_list[1]) )
                        A.place ( relx=0.15 , rely=0.3 )
                        B = tk.Button (self.root, command=B , text="B.%s" % (self.Update_list[2]) )
                        B.place ( relx=0.15 , rely=0.4 )
                        C = tk.Button ( self.root,command=C , text="C.%s" % (self.Update_list[3]) )
                        C.place ( relx=0.15 , rely=0.5 )
                        D = tk.Button ( self.root,command=D , text="D.%s" % (self.Update_list[4]) )
                        D.place ( relx=0.15 , rely=0.6 )
                        abort = tk.Button( self.root,command=abort , text="退出" )
                        abort.place(relx=0.15, rely=0.7)
                        jumpNum = tk.Entry(show=None)
                        jumpNum.place(relx=0.15, rely=0.8)

                    else:
                        def Check_result():
                            if vara.get ( ) == 1:
                                self.Check_List.append ( "A" )
                            if varb.get ( ) == 1:
                                self.Check_List.append ( "B" )
                            if varc.get ( ) == 1:
                                self.Check_List.append ( "C" )
                            if vard.get ( ) == 1:
                                self.Check_List.append ( "D" )
                            result = ''
                            for i in self.Check_List:
                                result += i
                            print ( result )
                            if result == Ti[5]:
                                self.dict['正确率'] += 2
                                Start(num+1)



                            else:
                                if self.dict['正确率'] > 0:
                                    self.dict['正确率'] -= 2
                                else:
                                    self.dict['正确率'] = 0
                                name = read ( 'data1.json' )
                                if self.True_count > name:
                                    with open ( "data1.json" , 'w' ) as f:
                                        json.dump ( self.True_count , f )
                                self.True_count = 0
                                self.Error.append ( self.Update_list )
                                tk.messagebox.showerror("Tip", "正确答案：%s" % self.True_Answer_type)
                                self.change -= 1
                                if self.change == 0:
                                    with open('../data.1','w') as f:
                                        f.write("No")
                                    chanage = tk.messagebox.askquestion ( '考试结束' , "你已经没有机会了，考试结束\n是否保存错题" )
                                    # print(chanage)
                                    if chanage == 'yes':
                                        self.tool.write_TiKu ( self.Error )
                                    self.Main ( )
                                    print ( "谢谢使用" )

                            self.Check_List = []
                        vara = tk.IntVar ( )
                        varb = tk.IntVar ( )
                        varc = tk.IntVar ( )
                        vard = tk.IntVar ( )
                        Ti = self.Update_list

                        A = tk.Checkbutton ( self.root,variable=vara , text="A.%s" % (Ti[1]) , onvalue=1 , offvalue=0 )
                        A.place ( relx=0.15 , rely=0.3 )
                        B = tk.Checkbutton ( self.root,variable=varb , text="B.%s" % (Ti[2]) , onvalue=1 , offvalue=0 )
                        B.place ( relx=0.15 , rely=0.4 )
                        C = tk.Checkbutton ( self.root,variable=varc , text="C.%s" % (Ti[3]) , onvalue=1 , offvalue=0 )
                        C.place ( relx=0.15 , rely=0.5 )
                        D = tk.Checkbutton ( self.root,variable=vard , text="D.%s" % (Ti[4]) , onvalue=1 , offvalue=0 )
                        D.place ( relx=0.15 , rely=0.6 )
                        result = tk.Button (self.root, text="提交" , command=Check_result )
                        result.place ( relx=0.15 , rely=0.7 )
                        abort = tk.Button( self.root,command=abort , text="退出" )
                        abort.place(relx=0.35, rely=0.7)


            else:
                tk.messagebox.showerror("Tip","没有这种类型.")


        Choose.place(relx=0.15,rely=0.4)

        Choose.current(0)
        Button_Goit = tk.Button(self.root,text="开始",command=Start(0))
        Button_Goit.place(relx=0.20,rely=0.62)
        Button_Back = tk.Button (self.root, text="回去" , command=self.Main )
        Button_Back.place ( relx=0.60 , rely=0.62 )


    def back( self ):

        def p(e):
            pass
        self.root.bind_all('<Control-s>',p)
        self.root.bind_all('<Control-S>',p)
        # msg.askquestion(title="返回", message="确定返回???") == "yes"
        if True:
            self.Main()
            self.mark = 0
            self.root.title("主菜单")
            self.root.resizable(False, False)
            self.__init__()



    def Main( self ):
        self.root.geometry("250x250")

        def get(event):

            choose = ListZhiShi.get(ListZhiShi.curselection())
            if choose == "考试":
                self.KaoShi()
            elif choose == "删号":
                self.delet()
            elif choose == "设置":
                self.setting()
            elif choose == "退出":
                result = msg.askquestion("Tip", "是否退出")
                if result == "yes":
                    self.root.destroy()
                else:
                    print(result)
            elif choose == "专题测试":
                self.Z_KaoShi()
            elif choose == "查看错题":
                self.Worre()
            elif choose == '账号更新':
                update()
        send("pass")

        menubar = tk.Menu ( self.root)


        menubar.add_command ( label="回到菜单" , command=self.back)

        listbox = tk.Menu(self.root,tearoff=0)

        self.root.config ( menu=menubar )

        PZ1 = tk.Label(self.root)


        PZ1.place(rely=0,relx=0,relwidth=1,relheight=1)
        ListZhiShi = tk.Listbox(self.root)
        List = ["考试","退出","设置",'查看错题',"专题测试"]
        title = tk.Label(self.root,text="更多考试")
        title.place(relx=0.7
                    ,rely=0.065
                    )
        def CuoTi():
            try:
                self.tk(self.tool.Cuoti())
            except:
                self.back()
                msg.showinfo('Tip',"没有错题!!")
        remove_use = tk.Button(self.root,text="错题考试\n错题做一遍",command=CuoTi)
        remove_use.place(relx=0.65,rely=0.25)

        try:
            with open('data2.json','r') as f:
                json.load(f)
        except Exception as e:
            List.append('账号更新')
            print(e)
        for i in List:
            ListZhiShi.insert(1,i)
        ListZhiShi.place(relx=0,rely=0)
        B = tk.Label(self.root,text="请选择你要参加的模式")
        B.place()
        ListZhiShi.bind("<Double-Button-1>",get)


    def tk(self,list):
        send("Yes")
        self.root.title("考试中..")
        self.indexa = 0
        self.FenShu = 0
        # 初始化
        self.root.geometry("400x400")
        if self.geometry == True:
            self.root.geometry("400x400")
            self.geometry = False
        if self.c == 0:
            msg.showinfo("Tips", "由于宽度限制，用户可以自行调整其宽度")
            self.c = 12
        self.start_time = time()
        def tk_one(hei):
            try:
                que = list[self.indexa][0]
            except IndexError:
                msg.showinfo("Tip",'考试完成,你的得分：%s'%(self.FenShu))
                self.showresult(self.FenShu,many=hei)
                return
            self.root.resizable(True,False)
            A  = list[self.indexa][1]
            B = list[self.indexa][2]
            C  = list[self.indexa][3]
            D = list[self.indexa][4]
            key = list[self.indexa][5]
            type2  = list[self.indexa][6]
            tk.Label().place(relwidth=1,relheight=1)
            # print("问题:%s,\nA:%s\nB:%s\nC:%s\nD:%s\nkey:%s"%(que,A,B,C,D,key))
            Que_text = tk.Text()
            Que_text.insert ( 'end' , que )
            Que_text.place(relx=0.15, rely=0.1, relwidth=0.7, relheight=0.2)
            def Check(answer):
                if answer == key:
                    if hei == 0:
                        msg.showinfo("Tip", "该sce文件题目数为0")
                        self.back()
                        return
                    self.dict["正确率"] += 5
                    self.save()
                    Fen = int(100/hei)
                    self.Add(mode=1,type_que=type2)
                    self.FenShu += Fen
                else:
                    try:
                        self.dict["正确率"] -= 5
                        self.save()
                        self.Error.append([que,A,B,C,D,key,answer,type2])
                    except IndexError:
                        msg.showinfo("Tip","考试完成")
                tk_one(hei)
            def A_F():
                self.indexa += 1
                Check("A")
            def B_F():
                self.indexa += 1
                Check("B")
            def C_F():
                self.indexa += 1
                Check("C")
            def D_F():
                self.indexa += 1
                Check("D")
            def Check_result():
                pass
            if len(key) == 1:
                A_B = tk.Button(self.root, command=A_F, text="A.%s" % (A))
                A_B.place(relx=0.15, rely=0.3)
                B_B = tk.Button(self.root, command=B_F, text="B.%s" % (B))
                B_B.place(relx=0.15, rely=0.4)
                C_B = tk.Button(self.root, command=C_F, text="C.%s" % (C))
                C_B.place(relx=0.15, rely=0.5)
                D_B = tk.Button(self.root, command=D_F, text="D.%s" % (D))
                D_B.place(relx=0.15, rely=0.6)
            else:
                def Check_result():
                    self.indexa += 1
                    if vara.get() == 1:
                        self.Check_List.append("A")
                    if varb.get() == 1:
                        self.Check_List.append("B")
                    if varc.get() == 1:
                        self.Check_List.append("C")
                    if vard.get() == 1:
                        self.Check_List.append("D")
                    result = ''
                    for i in self.Check_List:
                        result += i
                    Check(result)
                    # print(result)
                    self.Error.append([que,A,B,C,D,key,result,type2])
                    self.count += 1
                    if self.count == hei:
                        self.showresult(self.FenShu)
                    else:
                        tk_one(hei)

                vara = tk.IntVar()
                varb = tk.IntVar()
                varc = tk.IntVar()
                vard = tk.IntVar()
                self.Check_List = []
                A_B = tk.Checkbutton(self.root, variable=vara, text="A.%s" % (A), onvalue=1, offvalue=0)
                A_B.place(relx=0.15, rely=0.3)
                B_B = tk.Checkbutton(self.root, variable=varb, text="B.%s" % (B), onvalue=1, offvalue=0)
                B_B.place(relx=0.15, rely=0.4)
                C_B = tk.Checkbutton(self.root, variable=varc, text="C.%s" % (C), onvalue=1, offvalue=0)
                C_B.place(relx=0.15, rely=0.5)
                D_B = tk.Checkbutton(self.root, variable=vard, text="D.%s" % (D), onvalue=1, offvalue=0)
                D_B.place(relx=0.15, rely=0.6)
                result = tk.Button(self.root, text="提交", command=Check_result)
                result.place(relx=0.15, rely=0.7)
        tk_one(len(list))

    def showresult (self,mark,many="j"):
        '''考试结束的收尾'''
        if many == 'j':
            many = self.All

        # 时间加分
        send("pass")
        self.end_time = time()
        resu = self.end_time - self.start_time  # 答题总用时
        # print(resu)
        avg = int(resu / self.All)
        self.dict['速度'] += 10 - avg  # 得分
        print(self.dict)
        self.save()

        # 错题保存
        def S_B():

            self.tool.write_TiKu(self.Error)
            msg.showinfo("成功","成功添加了错题")
            self.back()

            self.hreat = 0
            self.mark = 0
            self.count = 0
            self.Error = []  # 存储错的题
            self.root.geometry ( "250x250" )



        PZ1 = tk.Label(self.root)
        PZ1.place(rely=0,relx=0,relwidth=1,relheight=1)
        self.geometry = True

        show = tk.Label(self.root,text="总分：%s分，您的得分:%s"%(100,int(mark)))
        show.place(relx=0.3,rely=0.3)
        Back = tk.Button(self.root,text="将错题添加到我的错题本上",command=S_B)
        Back.place(relx=0.35,rely=0.5)

        msg.showinfo("完成","做完题要记得查漏补缺哟")

    def save(self):
        '''保存雷达图数据'''
        with open("data2.json",'w') as f:
            json.dump(self.dict,f)


    def KaoShi( self ):
        list = self.tool.read_TiKu_type(modle="F")
        print(list)
        list = random.sample(list,self.All)
        self.tk(list)


def update():
    try:
        with open("data2.json",'r') as f:
            a = json.load(f)
        msg.showwarning("Tip","你的账号无需更新")
    except:
        with open("data2.json",'w') as f:
            json.dump({'速度':0,'正确率':0,
                       '单选':0,'多选':0,'判断':0},f)

        msg.showinfo("Tip","更新完毕")


class Login(): # 登陆对象

    def __init__(self):
        self.frist = True
        self.count = 0
        self.win = tk.Tk()
        self.win.title("登录")
        self.tool = Caozuo()
        # 读取base64转码后的数据，并设置压缩图标
        path = os.getcwd()
        ico = path + '/main.ico'
        self.win.wm_iconbitmap(ico)


        self.PZ1 = tk.Label()
        self.PZ1.place(rely=0,relx=0,relwidth=1,relheight=1)

        self.LabelUser = tk.Label(text="账号")
        self.LabelUser.place ( relx=0.15 , rely=0.1 )

        self.LabelPassWord = tk.Label ( text="密码" )
        self.LabelPassWord.place ( relx=0.15 , rely=0.3 )
        self.EntryUser = tk.Entry(show=None)
        self.EntryUser.place(relx=0.27,rely=0.1)

        self.EntryPas = tk.Entry(show="*")
        self.EntryPas.place(relx=0.27,rely=0.3)

        self.win.resizable(False,False)
        self.Wra = tk.Label(fg='red')
        self.ButtonLogon = tk.Button(text="注册",command=self.Logon,width=4)
        self.ButtonLogon.place(relx=0.27,rely=0.5)
        self.ButtonLogin = tk.Button(text="登录",command=self.Login,width=4)
        self.ButtonLogin.place(relx=0.5,rely=0.5)
        self.win.bind("<Return>",self.Reture)

        self.win.geometry("250x250")
        self.win.mainloop()
        # sys.exit()
        quit()
    def Reture(self,event
               ):
        if self.EntryPas.get() == None:
            msg.showwarning("Tip","密码没有输入")
        else:
            self.Login()
    def Login( self ):
        global UserNameFu
        with open("pass.json",'r') as F:
            st = json.load(F)
        User = self.EntryUser.get()
        if User not in st:
            msg.showinfo("错误","用户不存在")
        elif st[User] != self.EntryPas.get():
            msg.showinfo("错误","密码错误")
        else:
            self.win.destroy()# 结束时代
            UserNameFu = User
            C = Windows() # 开启新时代
    def Logon( self ):
        self.win.title ( "注册" )

        PZ1 = tk.Label ( )
        PZ1.place ( rely=0 , relx=0 , relwidth=1 , relheight=1 )
        LabelUser = tk.Label ( text="账号" )
        LabelUser.place ( relx=0.15 , rely=0.1 )
        LabelPassWord = tk.Label ( text="密码" )
        LabelPassWord.place ( relx=0.15 , rely=0.3 )
        EntryPas_logon = tk.Entry ( show="*" )
        EntryUser = tk.Entry ( show=None )
        EntryUser.place(relx=0.27 , rely=0.1 )
        EntryPas_logon.place ( relx=0.27 , rely=0.3 )

            # 保存窗口大小
            # 保存窗口大小

        def New ():
            self.win.title("注册")
            with open ( "pass.json" , 'r' ) as f:
                st = json.load ( f )

            if EntryPas_logon.get ( ) in st:
                msg.showinfo ("错误", "已有该用户，若忘记密码，你需要重新注册一个新的账号" )
            elif EntryPas_logon.get() == '' or EntryUser.get() == '':
                msg.showinfo("错误","没有输入密码账号")

            else:

                with open ( "pass.json" , "w" ) as f:
                    User = EntryUser.get()

                    passWord = EntryPas_logon.get()
                    st[User]=passWord
                    json.dump(st,f)
                try:
                    os.mkdir(User)
                except:
                    print('文件创建失败')
                with open("./%s/data1.json"%(User),'w') as f:
                    json.dump(0,f)
                with open("./%s/data2.json"%(User),'w') as f:
                    json.dump({'速度': 0, '正确率': 0,
                               '单选': 0, '多选': 0, '判断': 0}, f)
                    print("已为你创建了基本文件")
                    import openpyxl
                    wb = openpyxl.Workbook ( )
                    ws = wb.active
                    ws.title = "错题集"
                    wb.save("./%s/错题集.xlsx"%(User))

                    wb = load_workbook ( filename="./%s/错题集.xlsx"%(User) )
                    ws = wb["错题集"]
                    ws["A1"] = "题干"
                    ws["B1"] = 'A选项'
                    ws["C1"] = 'B选项'
                    ws["D1"] = "C选项"
                    ws["E1"] = "D选项"
                    ws["F1"] = "正确答案"
                    ws["G1"] = "你的答案"
                    msg.showinfo("成功","欢迎用户使用!!")
                    Sure.destroy()
                    EntryPas_logon.destroy()
                    EntryUser.destroy()
                    LabelPassWord.destroy()
                    LabelUser.destroy()
                    PZ1.destroy()
                    Return.destroy()
                    self.win.title ( "登录" )

        def Quit():
            self.win.title ( "登录" )
            Sure.destroy ( )
            EntryPas_logon.destroy ( )
            EntryUser.destroy ( )
            LabelPassWord.destroy ( )
            LabelUser.destroy ( )
            Return.destroy()
            PZ1.destroy ( )
        Sure = tk.Button(text="确定",command=New)
        Sure.place(relx=0.5,rely=0.5)

        Return = tk.Button(text="返回",command=Quit)
        Return.place(relx=0.2,rely=0.5)
if __name__ == '__main__':
    pid_F = os.getpid()
    with open("pid.data",'wb') as f:
        pickle.dump(pid_F,f)
    win = tk.Tk()
    with open("./DPI.data", 'wb') as f:
        pickle.dump([win.winfo_screenwidth(),win.winfo_screenheight()], f)  # [1920,1080]
    print("写入成功")
    win.destroy()
    send("No")
    poola = Pool(2)
    Main_Object = Login()
    send('close')
